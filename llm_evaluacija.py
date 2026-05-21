"""
Evaluacija jezičkih modela na srpskom jeziku
=============================================
Čita pitanja iz Excel fajla, šalje ih na 4 modela i upisuje odgovore nazad.

Potrebne biblioteke:
    pip install openpyxl openai anthropic google-generativeai requests tqdm

API ključevi se postavljaju u .env fajl (pogledaj README na dnu fajla).
"""

import os
import time
import logging
from pathlib import Path
from dotenv import load_dotenv
from tqdm import tqdm
import openpyxl

# ─── Učitaj .env fajl ────────────────────────────────────────────────────────
load_dotenv()

OPENAI_API_KEY    = os.getenv("OPENAI_API_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
MISTRAL_API_KEY   = os.getenv("MISTRAL_API_KEY")

# ─── Podešavanja ─────────────────────────────────────────────────────────────
EXCEL_FAJL     = "dataset_evaluacija_llm.xlsx"   # ← tvoj Excel fajl
DELAY_SEKUNDI  = 0.5   # pauza između poziva (da ne prekorači rate limit)
TEMPERATURA    = 0     # 0 = deterministic, bitno za konzistentnost
MAX_TOKENA     = 500   # maksimalna dužina odgovora

# Koji modeli se testiraju (True/False za svaki)
TESTIRAJ = {
    "gpt4":    True,
    "claude":  True,
    "mistral": True,
}

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("evaluacija.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════════════
# KLIJENTI ZA SVAKI MODEL
# ═════════════════════════════════════════════════════════════════════════════

def pitaj_gpt4(pitanje: str) -> str:
    """Šalje pitanje GPT-4o modelu."""
    try:
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4o",
            temperature=TEMPERATURA,
            max_tokens=MAX_TOKENA,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Ti si asistent koji odgovara na pitanja na srpskom jeziku. "
                        "Odgovori kratko i precizno. "
                        "Ako nisi siguran ili pitanje nije smisleno, reci to jasno."
                    )
                },
                {"role": "user", "content": pitanje}
            ]
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        log.error(f"GPT-4 greška: {e}")
        return f"GREŠKA: {e}"


def pitaj_claude(pitanje: str) -> str:
    """Šalje pitanje Claude modelu."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=MAX_TOKENA,
            temperature=TEMPERATURA,
            system=(
                "Ti si asistent koji odgovara na pitanja na srpskom jeziku. "
                "Odgovori kratko i precizno. "
                "Ako nisi siguran ili pitanje nije smisleno, reci to jasno."
            ),
            messages=[{"role": "user", "content": pitanje}]
        )
        return response.content[0].text.strip()
    except Exception as e:
        log.error(f"Claude greška: {e}")
        return f"GREŠKA: {e}"



def pitaj_mistral(pitanje: str, pokusaj: int = 0) -> str:
    """Šalje pitanje Mistral modelu via API sa automatskim retry."""
    try:
        import requests
        headers = {
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "mistral-small-latest",
            "temperature": TEMPERATURA,
            "max_tokens": MAX_TOKENA,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Ti si asistent koji odgovara na pitanja na srpskom jeziku. "
                        "Odgovori kratko i precizno. "
                        "Ako nisi siguran ili pitanje nije smisleno, reci to jasno."
                    )
                },
                {"role": "user", "content": pitanje}
            ]
        }
        r = requests.post(
            "https://api.mistral.ai/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=30
        )
        if r.status_code == 429 and pokusaj < 4:
            cekanje = 30 * (pokusaj + 1)
            log.warning(f"Mistral rate limit — cekam {cekanje}s (pokusaj {pokusaj+1}/4)...")
            time.sleep(cekanje)
            return pitaj_mistral(pitanje, pokusaj + 1)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        log.error(f"Mistral greška: {e}")
        return f"GREŠKA: {e}"


# ═════════════════════════════════════════════════════════════════════════════
# KONFIGURACIJA SHEET-OVA
# Definišemo za svaki sheet: kolonu sa pitanjem i kolone za odgovore modela
# ═════════════════════════════════════════════════════════════════════════════

SHEET_CONFIG = {
    # Raspored Tacnost :
    # A=ID, B=Kategorija, C=Pitanje, D=Tacan odgovor, E=Izvor, F=Tezina,
    # G=Odg.GPT, H=Odg.Claude,  J=Odg.Mistral, K=Tacno?, L=Napomena
    "Tačnost": {
        "pitanje_col": 3,
        "pocetak_reda": 3,
        "modeli": {
            "gpt4":    7,
            "claude":  8,
           
            "mistral": 9,
        }
    },
    # Raspored Halucinacije:
    # A=ID, B=Tip zamke, C=Pitanje, D=Ocekivano, E=Crvene zastavice,
    # F=Odg.GPT, G=Odg.Claude,  I=Odg.Mistral
    "Halucinacije": {
        "pitanje_col": 3,
        "pocetak_reda": 3,
        "modeli": {
            "gpt4":    6,
            "claude":  7,
            
            "mistral": 8,
        }
    },
    # Raspored Konzistentnost:
    # A=Grupa ID, B=Varijanta, C=Tip formulacije, D=Pitanje, E=Tacan odgovor,
    # F=Odg.GPT, H=Odg.Gemini, I=Odg.Mistral
    "Konzistentnost": {
        "pitanje_col": 4,
        "pocetak_reda": 3,
        "modeli": {
            "gpt4":    6,
            "claude":  7,
           
            "mistral": 8,
        }
    },
}

# Mapiranje naziva modela na funkcije
MODEL_FUNKCIJE = {
    "gpt4":    pitaj_gpt4,
    "claude":  pitaj_claude,
   
    "mistral": pitaj_mistral,
}

MODEL_NAZIVI = {
    "gpt4":    "GPT-4o",
    "claude":  "Claude Sonnet",
 
    "mistral": "Mistral Large",
}


# ═════════════════════════════════════════════════════════════════════════════
# GLAVNA LOGIKA
# ═════════════════════════════════════════════════════════════════════════════

def dodaj_zaglavlja_modela(ws, config: dict):
    """Dodaje zaglavlja kolona za modele koji nisu bili u originalnom sheetu."""
    from openpyxl.styles import Font, PatternFill, Alignment

    for naziv_modela, col in config["modeli"].items():
        if not TESTIRAJ.get(naziv_modela):
            continue
        cell = ws.cell(row=2, column=col)
        if not cell.value:  # samo ako je prazno
            cell.value = f"Odg. {MODEL_NAZIVI[naziv_modela]}"
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color="444444")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 35


def procesiraj_sheet(wb, sheet_naziv: str, config: dict, izlaz: str = "rezultati.xlsx"):
    """Čita pitanja iz sheeta i upisuje odgovore svih modela."""
    ws = wb[sheet_naziv]
    dodaj_zaglavlja_modela(ws, config)

    pitanje_col   = config["pitanje_col"]
    pocetak_reda  = config["pocetak_reda"]
    max_red       = ws.max_row

    # Broji neprazna pitanja
    pitanja = []
    for red in range(pocetak_reda, max_red + 1):
        pitanje = ws.cell(row=red, column=pitanje_col).value
        if pitanje and str(pitanje).strip():
            pitanja.append((red, str(pitanje).strip()))

    if not pitanja:
        log.warning(f"Sheet '{sheet_naziv}': nema pitanja!")
        return

    log.info(f"\n{'═'*60}")
    log.info(f"Sheet: {sheet_naziv} — {len(pitanja)} pitanja")
    log.info(f"{'═'*60}")

    for naziv_modela, col_odgovora in config["modeli"].items():
        if not TESTIRAJ.get(naziv_modela):
            continue

        # Provera da li model postoji u mapi funkcija
        if naziv_modela not in MODEL_FUNKCIJE:
            log.error(f"Model '{naziv_modela}' nije definisan u MODEL_FUNKCIJE! Preskačem.")
            log.error(f"Dostupni modeli: {list(MODEL_FUNKCIJE.keys())}")
            continue

        funkcija = MODEL_FUNKCIJE[naziv_modela]
        log.info(f"\n  ▶ Model: {MODEL_NAZIVI[naziv_modela]}")

        for red, pitanje in tqdm(pitanja, desc=f"  {MODEL_NAZIVI[naziv_modela]}"):
            # Preskoči ako već ima odgovor
            postojeci = ws.cell(row=red, column=col_odgovora).value
            if postojeci and str(postojeci).strip() and not str(postojeci).startswith("GREŠKA"):
                continue

            odgovor = funkcija(pitanje)
            ws.cell(row=red, column=col_odgovora).value = odgovor

            # Ispiši odgovor u terminal
            kratko_pitanje = pitanje[:60] + "..." if len(pitanje) > 60 else pitanje
            kratak_odgovor = odgovor[:80] + "..." if len(odgovor) > 80 else odgovor
            log.info(f"    [Red {red}] P: {kratko_pitanje}")
            log.info(f"            O: {kratak_odgovor}")

            # Sačuvaj posle svakog pitanja (zaštita od prekida)
            wb.save(izlaz)

            time.sleep(DELAY_SEKUNDI)

        log.info(f"  ✓ Gotovo: {MODEL_NAZIVI[naziv_modela]}")


def glavna_funkcija():
    """Učitava Excel, procesira sve sheetove, čuva rezultate."""
    putanja = Path(EXCEL_FAJL)
    if not putanja.exists():
        log.error(f"Fajl '{EXCEL_FAJL}' nije pronađen!")
        return

    log.info(f"Učitavam: {EXCEL_FAJL}")
    wb = openpyxl.load_workbook(putanja)

    # ── Dijagnostika: ispiši zaglavlja svih sheetova ──────────────────────
    log.info("── Zaglavlja sheetova (red 2) ──")
    for naziv in wb.sheetnames:
        ws_d = wb[naziv]
        zaglavlja = {
            col: ws_d.cell(row=2, column=col).value
            for col in range(1, ws_d.max_column + 1)
            if ws_d.cell(row=2, column=col).value
        }
        log.info(f"  [{naziv}] {zaglavlja}")
    log.info("────────────────────────────────")

    # Definisi izlazni fajl pre petlje (cuvamo posle svakog pitanja)
    izlaz = putanja.stem + "_sa_odgovorima.xlsx"
    log.info(f"Odgovori ce biti sacuvani u: {izlaz}")

    for sheet_naziv, config in SHEET_CONFIG.items():
        if sheet_naziv in wb.sheetnames:
            procesiraj_sheet(wb, sheet_naziv, config, izlaz)
        else:
            log.warning(f"Sheet '{sheet_naziv}' ne postoji u fajlu!")

    wb.save(izlaz)
    log.info(f"\n✅ Svi rezultati sacuvani u: {izlaz}")


if __name__ == "__main__":
    glavna_funkcija()


# ═════════════════════════════════════════════════════════════════════════════
# README — HOW TO RUN
# ═════════════════════════════════════════════════════════════════════════════
"""
INSTALACIJA:
    pip install openpyxl openai anthropic google-generativeai requests tqdm python-dotenv

KREIRANJE .env FAJLA:
    Napravi fajl '.env' u istom folderu kao i ovaj skript, sa sadržajem:

        OPENAI_API_KEY=sk-...
        ANTHROPIC_API_KEY=sk-ant-...
        GEMINI_API_KEY=AIza...
        MISTRAL_API_KEY=...

    NIKAD ne deli .env fajl niti ga stavljaj na GitHub!

POKRETANJE:
    python llm_evaluacija.py

NAPOMENE:
    - Skript automatski preskače pitanja koja već imaju odgovor (resume podrška)
    - Svi odgovori se loguju u 'evaluacija.log'
    - Izlazni fajl: 'dataset_evaluacija_llm_sa_odgovorima.xlsx'
    - Temperatura = 0 osigurava reproduktivnost rezultata
"""
